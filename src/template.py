import io

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

ENVIOS_EJEMPLO = pd.DataFrame(
    [
        {"CD": "BK31", "SKU": 1001, "Descripción": "Ron Ejemplo 750ml 1X1", "Cajas Teóricas": 10, "Unidades": 120},
        {"CD": "BK31", "SKU": 1002, "Descripción": "Yogurt Ejemplo 1L", "Cajas Teóricas": 25, "Unidades": 300},
        {"CD": "BK34", "SKU": 1003, "Descripción": "Cigarro Ejemplo Cajetilla", "Cajas Teóricas": 3.5, "Unidades": 70},
        {"CD": "BK34", "SKU": 1001, "Descripción": "Ron Ejemplo 750ml 1X1", "Cajas Teóricas": 15, "Unidades": 180},
        {"CD": "BK34", "SKU": 1004, "Descripción": "Four Loko Ejemplo 473ml 6x1", "Cajas Teóricas": 8, "Unidades": 48},
    ]
)

MAESTRO_EJEMPLO = pd.DataFrame(
    [
        {"SKU": 1001, "Categoría": "Licores", "Subcategoría": "", "Unidades por caja": 12, "Cajas por cama": 20, "Camas por PH": 6, "Cajas por PH": 120},
        {"SKU": 1002, "Categoría": "Lácteos", "Subcategoría": "", "Unidades por caja": 12, "Cajas por cama": 15, "Camas por PH": 7, "Cajas por PH": 105},
        {"SKU": 1003, "Categoría": "Cigarros", "Subcategoría": "", "Unidades por caja": 20, "Cajas por cama": 50, "Camas por PH": 3, "Cajas por PH": 150},
        {"SKU": 1004, "Categoría": "NABs", "Subcategoría": "RTD", "Unidades por caja": 6, "Cajas por cama": 30, "Camas por PH": 4, "Cajas por PH": 120},
    ]
)

UMA_EJEMPLO = pd.DataFrame(
    [
        {"SKU": 1001, "Largo de caja": 30, "Ancho de caja": 20, "Alto de caja": 25, "Peso bruto por unidad": 0.90},
        {"SKU": 1002, "Largo de caja": 25, "Ancho de caja": 25, "Alto de caja": 15, "Peso bruto por unidad": 1.00},
        {"SKU": 1003, "Largo de caja": 35, "Ancho de caja": 25, "Alto de caja": 55, "Peso bruto por unidad": 0.05},
        {"SKU": 1004, "Largo de caja": 22, "Ancho de caja": 15, "Alto de caja": 20, "Peso bruto por unidad": 0.40},
    ]
)

PALLETS_OBJETIVO_EJEMPLO = pd.DataFrame(
    [
        {"CD": "BK34", "Pallets_Objetivo": 1},
    ]
)

INSTRUCCIONES = pd.DataFrame(
    [
        {"Hoja": "Envios_Julio", "Columna": "CD", "Qué va aquí": "Código del Centro de Distribución destino (texto), ej. BK31."},
        {"Hoja": "Envios_Julio", "Columna": "SKU", "Qué va aquí": "Identificador del producto. Debe existir también en Maestro_SKUs y en UMA."},
        {"Hoja": "Envios_Julio", "Columna": "Descripción", "Qué va aquí": "Nombre o descripción del producto."},
        {"Hoja": "Envios_Julio", "Columna": "Cajas Teóricas", "Qué va aquí": "Demanda oficial en cajas. Solo puede traer decimales si la Categoría del SKU es Cigarros (se redondea hacia arriba)."},
        {"Hoja": "Envios_Julio", "Columna": "Unidades", "Qué va aquí": "Demanda en unidades sueltas (informativo)."},
        {"Hoja": "Maestro_SKUs", "Columna": "SKU", "Qué va aquí": "Debe coincidir exactamente con el SKU usado en Envios_Julio y UMA."},
        {"Hoja": "Maestro_SKUs", "Columna": "Categoría", "Qué va aquí": "Una de: Licores, Lácteos, Aseo, Importados, Merch, NABs, Comestibles, Cigarros. Cualquier otro valor queda marcado como 'no clasificada' y no se apila automáticamente."},
        {"Hoja": "Maestro_SKUs", "Columna": "Subcategoría", "Qué va aquí": "Opcional. Escribir 'RTD' o 'Energizante' en los SKUs de esas subcategorías (incluye Four Loko) para que se traten como frágiles -nunca se les pone nada encima, van al final/arriba del pallet. Dejar vacío en el resto."},
        {"Hoja": "Maestro_SKUs", "Columna": "Unidades por caja", "Qué va aquí": "Cuántas unidades sueltas trae una caja de este SKU."},
        {"Hoja": "Maestro_SKUs", "Columna": "Cajas por cama", "Qué va aquí": "Cuántas cajas de este SKU caben en una cama (capa) del pallet. Si se deja vacío o en 0, se calcula automáticamente desde las dimensiones de UMA."},
        {"Hoja": "Maestro_SKUs", "Columna": "Camas por PH", "Qué va aquí": "Cuántas camas conforman un pallet homogéneo completo de este SKU."},
        {"Hoja": "Maestro_SKUs", "Columna": "Cajas por PH", "Qué va aquí": "Cuántas cajas conforman un pallet homogéneo completo de este SKU."},
        {"Hoja": "UMA", "Columna": "SKU", "Qué va aquí": "Debe coincidir exactamente con el SKU usado en Envios_Julio y Maestro_SKUs."},
        {"Hoja": "UMA", "Columna": "Largo de caja / Ancho de caja", "Qué va aquí": "Dimensiones de la base de la caja, en centímetros. El sistema prueba automáticamente ambas rotaciones sobre el pallet de 120x100 cm."},
        {"Hoja": "UMA", "Columna": "Alto de caja", "Qué va aquí": "Altura de la caja en centímetros. Máximo permitido: 180.08 cm."},
        {"Hoja": "UMA", "Columna": "Peso bruto por unidad", "Qué va aquí": "Peso en kg de una unidad suelta (se multiplica por 'Unidades por caja' de Maestro_SKUs para obtener el peso de la caja)."},
        {"Hoja": "Pallets_Objetivo", "Columna": "(hoja opcional)", "Qué va aquí": "Borrar la hoja entera si no se necesita -sin ella, el sistema abre tantos pallets como haga falta (comportamiento de siempre). Si se incluye un CD acá, TODA su demanda se reparte en exactamente esa cantidad de pallets -viene de la planificación externa, el sistema no la calcula. OJO: puede tardar varios minutos por pallet listado (motor exacto + ruina-y-reconstrucción)."},
        {"Hoja": "Pallets_Objetivo", "Columna": "CD", "Qué va aquí": "Código del CD, igual que en Envios_Julio."},
        {"Hoja": "Pallets_Objetivo", "Columna": "Pallets_Objetivo", "Qué va aquí": "Cantidad FIJA de pallets para ese CD (entero, ej. 5)."},
    ]
)


def _autoajustar_columnas(hoja, df: pd.DataFrame) -> None:
    for i, columna in enumerate(df.columns, start=1):
        ancho = max(len(str(columna)), df[columna].astype(str).map(len).max() if not df.empty else 0) + 4
        hoja.column_dimensions[hoja.cell(row=1, column=i).column_letter].width = min(ancho, 60)


def construir_template(ruta_o_buffer=None):
    destino = ruta_o_buffer if ruta_o_buffer is not None else io.BytesIO()
    with pd.ExcelWriter(destino, engine="openpyxl") as writer:
        INSTRUCCIONES.to_excel(writer, sheet_name="Instrucciones", index=False)
        ENVIOS_EJEMPLO.to_excel(writer, sheet_name="Envios_Julio", index=False)
        MAESTRO_EJEMPLO.to_excel(writer, sheet_name="Maestro_SKUs", index=False)
        UMA_EJEMPLO.to_excel(writer, sheet_name="UMA", index=False)
        PALLETS_OBJETIVO_EJEMPLO.to_excel(writer, sheet_name="Pallets_Objetivo", index=False)

        libro = writer.book
        encabezado_relleno = PatternFill(start_color="2A78D6", end_color="2A78D6", fill_type="solid")
        encabezado_fuente = Font(color="FFFFFF", bold=True)

        for nombre_hoja, df in (
            ("Instrucciones", INSTRUCCIONES),
            ("Envios_Julio", ENVIOS_EJEMPLO),
            ("Maestro_SKUs", MAESTRO_EJEMPLO),
            ("UMA", UMA_EJEMPLO),
            ("Pallets_Objetivo", PALLETS_OBJETIVO_EJEMPLO),
        ):
            hoja = libro[nombre_hoja]
            for celda in hoja[1]:
                celda.fill = encabezado_relleno
                celda.font = encabezado_fuente
                celda.alignment = Alignment(wrap_text=True, vertical="center")
            _autoajustar_columnas(hoja, df)

    if ruta_o_buffer is None:
        destino.seek(0)
    return destino
