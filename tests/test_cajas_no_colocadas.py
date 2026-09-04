"""[pedido explícito del usuario: "quiero que las cajas de cada cd que no
logre cubicar la ponga en una hoja de resultado... igual que el detalle
por cd y sku y cantidad"] `armar_pallets_bloques` ya calculaba
`sin_colocar` (SKU -> cantidad) pero se perdía en silencio entre el motor
de armado y el Excel final -`_palletv5_a_pallet` (pipeline_sku_bloque.py)
solo copiaba `torres`/`cajas_bat` de cada `PalletV5`, nunca `metadata`.
Estos tests fijan el contrato: la demanda sin colocar tiene que llegar
hasta `ResultadoPipeline.cajas_no_colocadas_df` y desde ahí a la hoja
`Cajas_No_Colocadas` del Excel exportado."""
import openpyxl

from src.exportar import construir_cajas_no_colocadas_df, exportar_workbook
from src.pipeline import ejecutar_pipeline


def test_sin_faltantes_el_df_queda_vacio(dataset_factory):
    """[no-regresión] Demanda que entra sin problema -el DF debe quedar
    vacío (o `None`), nunca inventar filas."""
    envios, maestro, uma = dataset_factory(envios_overrides=[{"sku": 1, "cajas": 5}])
    resultado = ejecutar_pipeline(envios, maestro, uma)
    assert resultado.cajas_no_colocadas_df is None or resultado.cajas_no_colocadas_df.empty


def test_pallets_objetivo_insuficiente_reporta_el_faltante_por_cd_y_sku(dataset_factory):
    """[caso real que motivó el pedido] Un CD con `Pallets_Objetivo`
    fijo demasiado chico para toda su demanda -la demanda que no entra
    debe aparecer con su CD y SKU exactos, no perderse."""
    envios, maestro, uma = dataset_factory(
        envios_overrides=[{"cd": "BK31", "sku": 1, "cajas": 500}],
        maestro_overrides=[{"sku": 1, "categoria": "Licores", "cajas_por_ph": 50}],
        uma_overrides=[{"sku": 1}],
    )
    resultado = ejecutar_pipeline(envios, maestro, uma, pallets_objetivo_por_cd={"BK31": 1})
    df = resultado.cajas_no_colocadas_df
    assert df is not None and not df.empty
    fila = df[df["SKU"] == "1"].iloc[0]
    assert fila["CD"] == "BK31"
    assert fila["Cajas_No_Colocadas"] > 0


def test_hoja_cajas_no_colocadas_aparece_en_el_excel_solo_si_hay_faltantes(dataset_factory, tmp_path):
    envios, maestro, uma = dataset_factory(
        envios_overrides=[{"cd": "BK31", "sku": 1, "cajas": 500}],
        maestro_overrides=[{"sku": 1, "categoria": "Licores", "cajas_por_ph": 50}],
        uma_overrides=[{"sku": 1}],
    )
    resultado = ejecutar_pipeline(envios, maestro, uma, pallets_objetivo_por_cd={"BK31": 1})
    ruta = tmp_path / "salida.xlsx"
    exportar_workbook(resultado, ruta)
    libro = openpyxl.load_workbook(ruta)
    assert "Cajas_No_Colocadas" in libro.sheetnames

    envios_ok, maestro_ok, uma_ok = dataset_factory(envios_overrides=[{"sku": 1, "cajas": 5}])
    resultado_ok = ejecutar_pipeline(envios_ok, maestro_ok, uma_ok)
    ruta_ok = tmp_path / "salida_ok.xlsx"
    exportar_workbook(resultado_ok, ruta_ok)
    libro_ok = openpyxl.load_workbook(ruta_ok)
    assert "Cajas_No_Colocadas" not in libro_ok.sheetnames


def test_construir_cajas_no_colocadas_df_agrupa_por_cd_sku(dataset_factory):
    """`construir_cajas_no_colocadas_df` en aislado -reusa `info_sku` para
    traer descripción/categoría, y agrupa si el mismo SKU quedara marcado
    sin colocar en más de un `PalletV5` del mismo CD (no debería pasar hoy,
    pero la función no debe duplicar la fila si pasara)."""
    from models import PalletV5

    p1 = PalletV5(id="P1", cd="BK31", metadata={"sin_colocar": {"1": 3}})
    p2 = PalletV5(id="P2", cd="BK31", metadata={"sin_colocar": {"1": 2}})
    p3 = PalletV5(id="P3", cd="BK99", metadata={})
    df = construir_cajas_no_colocadas_df([p1, p2, p3], info_sku={"1": {"descripcion": "X", "categoria": "Licores"}})
    assert len(df) == 1
    assert df.iloc[0]["CD"] == "BK31"
    assert df.iloc[0]["Cajas_No_Colocadas"] == 5
    assert df.iloc[0]["Descripcion"] == "X"
